"""Generacion del reporte .xlsx con las cargas registradas."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Exercise, Routine, RoutineExercise, SetLog, User

HEADERS = [
    "Fecha",
    "Bateria",
    "Jugador",
    "Ejercicio",
    "Categoria",
    "Serie",
    "Kg",
    "Reps hechas",
    "Reps objetivo",
    "Registrado",
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _fetch_rows(
    db: Session,
    date_from: date | None,
    date_to: date | None,
    player_id: int | None,
) -> list[tuple]:
    stmt = (
        select(
            Routine.session_date,
            Routine.name,
            User.name,
            Exercise.name,
            Exercise.category,
            SetLog.set_number,
            SetLog.load_kg,
            SetLog.reps,
            RoutineExercise.target_reps,
            SetLog.updated_at,
        )
        .join(User, User.id == SetLog.user_id)
        .join(RoutineExercise, RoutineExercise.id == SetLog.routine_exercise_id)
        .join(Routine, Routine.id == RoutineExercise.routine_id)
        .join(Exercise, Exercise.id == RoutineExercise.exercise_id)
        .order_by(Routine.session_date, User.name, RoutineExercise.position, SetLog.set_number)
    )
    if date_from:
        stmt = stmt.where(Routine.session_date >= date_from)
    if date_to:
        stmt = stmt.where(Routine.session_date <= date_to)
    if player_id:
        stmt = stmt.where(SetLog.user_id == player_id)

    return list(db.execute(stmt).all())


def _autosize(sheet) -> None:
    for index, header in enumerate(HEADERS, start=1):
        longest = max(
            [len(header)] + [len(str(cell.value or "")) for cell in sheet[get_column_letter(index)][1:]]
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(longest + 3, 40)


def build_report(
    db: Session,
    date_from: date | None = None,
    date_to: date | None = None,
    player_id: int | None = None,
) -> BytesIO:
    workbook = Workbook()
    detail = workbook.active
    detail.title = "Cargas"
    detail.append(HEADERS)
    for cell in detail[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    rows = _fetch_rows(db, date_from, date_to, player_id)
    for row in rows:
        detail.append(
            [
                row[0],
                row[1],
                row[2],
                row[3],
                row[4] or "",
                row[5],
                float(row[6]),
                row[7],
                row[8],
                row[9].replace(tzinfo=None) if row[9] else None,
            ]
        )

    for row_cells in detail.iter_rows(min_row=2, min_col=1, max_col=1):
        row_cells[0].number_format = "DD/MM/YYYY"
    for row_cells in detail.iter_rows(min_row=2, min_col=10, max_col=10):
        row_cells[0].number_format = "DD/MM/YYYY HH:MM"

    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(detail.max_row, 1)}"
    _autosize(detail)

    _add_summary_sheet(workbook, rows)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _add_summary_sheet(workbook: Workbook, rows: list[tuple]) -> None:
    """Volumen total (kg x reps) y carga maxima por jugador y ejercicio."""
    summary = workbook.create_sheet("Resumen")
    headers = ["Jugador", "Ejercicio", "Series", "Kg maximo", "Kg medio", "Volumen (kg x reps)"]
    summary.append(headers)
    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    aggregated: dict[tuple[str, str], dict] = {}
    for row in rows:
        key = (row[2], row[3])
        load = float(row[6])
        reps = row[7] or 0
        entry = aggregated.setdefault(key, {"sets": 0, "max": 0.0, "total": 0.0, "volume": 0.0})
        entry["sets"] += 1
        entry["max"] = max(entry["max"], load)
        entry["total"] += load
        entry["volume"] += load * reps

    for (player, exercise), entry in sorted(aggregated.items()):
        summary.append(
            [
                player,
                exercise,
                entry["sets"],
                round(entry["max"], 2),
                round(entry["total"] / entry["sets"], 2) if entry["sets"] else 0,
                round(entry["volume"], 2),
            ]
        )

    summary.freeze_panes = "A2"
    for index, header in enumerate(headers, start=1):
        summary.column_dimensions[get_column_letter(index)].width = max(len(header) + 4, 16)
