from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_export_excel_devuelve_un_xlsx_con_dos_hojas(client: TestClient, coach_headers: dict):
    response = client.get("/api/export/excel", headers=coach_headers)

    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MIME
    assert "attachment" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Cargas", "Resumen"]
    assert workbook["Cargas"][1][0].value == "Fecha"
    assert workbook["Cargas"][1][6].value == "Kg"


def test_export_filtrado_por_fechas_sin_datos(client: TestClient, coach_headers: dict):
    response = client.get(
        "/api/export/excel", params={"date_from": "1990-01-01", "date_to": "1990-01-31"}, headers=coach_headers
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Cargas"].max_row == 1
